import os
import json
from pathlib import Path
import openpyxl

# Path ke folder data
DATA_FOLDER = Path(__file__).parent / "data"
OUTPUT_FILE = Path(__file__).parent / "data.json"

# Nama kolom yang kita cari di Excel
COLUMN_NAMES = [
    'kodeItem', 'barcode', 'namaItem', 'hargaPokok', 
    'kodeBarang', 'hargaJual', 'stok'
]

HEADER_PATTERNS = {
    'kodeItem': [('kode', 'item')],
    'barcode': [('barcode',)],
    'namaItem': [('nama', 'item')],
    'hargaPokok': [('harga', 'pokok')],
    'kodeBarang': [('kode', 'barang'), ('kode', 'harga')],
    'hargaJual': [('harga', 'jual')],
    'stok': [('stok',)],
}

NUMERIC_COLUMNS = {'hargaPokok', 'hargaJual', 'stok'}


def normalize_header(value):
    """Normalisasi nama header supaya mudah dicocokkan."""
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().replace('_', ' ').split())


def resolve_header(normalized):
    """Pilih nama kolom standar dari variasi header Excel."""
    for canonical, patterns in HEADER_PATTERNS.items():
        for pattern in patterns:
            if all(token in normalized for token in pattern):
                return canonical
    return None


def to_number(value):
    """Konversi angka dari Excel ke integer yang aman untuk JSON."""
    if value in (None, ""):
        return 0

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip().replace('.', '').replace(',', '')
    if not text:
        return 0

    try:
        return int(float(text))
    except ValueError:
        return 0


def to_text(value):
    """Pertahankan teks seperti kode item dan barcode tanpa suffix .0."""
    if value in (None, ""):
        return ""

    if isinstance(value, bool):
        return str(int(value))

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)

    return str(value).strip()

def find_header_row_and_columns(ws, column_names):
    """Cari baris header dan index kolom dari beberapa baris pertama."""
    max_scan_row = min(ws.max_row, 15)
    best_row = None
    best_map = {}

    for row_idx in range(1, max_scan_row + 1):
        column_map = {}
        for col_idx, cell in enumerate(ws[row_idx], 1):
            normalized = normalize_header(cell.value)
            if not normalized:
                continue

            canonical = resolve_header(normalized)
            if canonical in column_names and canonical not in column_map:
                column_map[canonical] = col_idx

        if len(column_map) > len(best_map):
            best_row = row_idx
            best_map = column_map

    required_headers = {'kodeItem', 'namaItem', 'hargaPokok', 'hargaJual', 'stok'}
    if best_row and len(required_headers.intersection(best_map.keys())) >= 4:
        return best_row, best_map

    return None, {}

def read_excel_file(file_path):
    """Membaca file Excel dan mengembalikan list of items"""
    items = []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Cari baris header dan index kolom
        header_row, column_map = find_header_row_and_columns(ws, COLUMN_NAMES)

        if not column_map or header_row is None:
            print(f"Tidak ada header yang cocok di {file_path.name}")
            return items

        # Baca data mulai setelah baris header
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            # Cek apakah row kosong
            if all(cell.value is None for cell in row):
                continue

            item = {}
            try:
                for col_name in COLUMN_NAMES:
                    col_idx = column_map.get(col_name)
                    cell_value = row[col_idx - 1].value if col_idx else None

                    if col_name in NUMERIC_COLUMNS:
                        item[col_name] = to_number(cell_value)
                    else:
                        item[col_name] = to_text(cell_value)

                # Hanya tambah jika ada data minimal
                if item.get('namaItem') and any(item.values()):
                    items.append(item)
            except Exception as e:
                print(f"  ⚠️  Error di row {row_idx}: {e}")
                continue

        wb.close()
    except Exception as e:
        print(f"Error membaca {file_path}: {e}")

    return items

def build_json_structure(data_folder):
    """Build struktur JSON berdasarkan folder hierarchy"""
    result = {}
    
    for item in sorted(os.listdir(data_folder)):
        item_path = data_folder / item
        
        if os.path.isdir(item_path):
            # Ini adalah folder lokasi (Lantai 1, Lorong A, dll)
            print(f"\nProses folder: {item}")
            result[item] = process_folder(item_path)
    
    return result

def process_folder(folder_path):
    """Proses folder dan return struktur data (bisa array atau dict)"""
    has_excel = False
    has_subfolder = False
    items = []
    subfolders = {}

    for item in sorted(os.listdir(folder_path)):
        item_path = folder_path / item

        if os.path.isdir(item_path):
            # Ada subfolder
            has_subfolder = True
            print(f"  Subfolder: {item}")
            subfolders[item] = process_folder(item_path)
        elif item.endswith('.xlsx'):
            # Ada file Excel
            has_excel = True
            print(f"  File: {item}")
            file_items = read_excel_file(item_path)
            if has_subfolder:
                subfolders[item_path.stem] = file_items
            else:
                items.extend(file_items)
            print(f"    {len(file_items)} item terbaca")

    # Return struktur yang sesuai
    if has_subfolder:
        # Folder campuran disimpan sebagai object agar tiap file tetap punya label
        return subfolders

    # Hanya ada file Excel, return array gabungan
    return items

def main():
    print("=" * 60)
    print("KONVERSI EXCEL KE JSON")
    print("=" * 60)
    
    if not DATA_FOLDER.exists():
        print(f"Folder {DATA_FOLDER} tidak ditemukan!")
        return
    
    print(f"Folder data: {DATA_FOLDER}\n")
    
    # Build struktur
    data = build_json_structure(DATA_FOLDER)
    
    # Simpan ke JSON
    try:
        with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"\n{'=' * 60}")
        print(f"BERHASIL! File disimpan ke: {OUTPUT_FILE}")
        print(f"{'=' * 60}")
        
        # Summary
        total_items = count_total_items(data)
        print(f"Total item yang terbaca: {total_items}")
        
    except Exception as e:
        print(f"Error menyimpan JSON: {e}")

def count_total_items(data, count=0):
    """Hitung total items di seluruh struktur"""
    for value in data.values():
        if isinstance(value, list):
            count += len(value)
        elif isinstance(value, dict):
            count += count_total_items(value, 0)
    return count

if __name__ == "__main__":
    main()
