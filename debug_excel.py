import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('data/Lantai 1/Aki.xlsx', data_only=True)
    ws = wb.active
    print(f'Total rows: {ws.max_row}', file=sys.stderr)
    print(f'Total cols: {ws.max_column}', file=sys.stderr)
    print('Headers:', file=sys.stderr)
    for i, cell in enumerate(ws[1], 1):
        print(f'  Col {i}: {cell.value}', file=sys.stderr)
    
    if ws.max_row > 1:
        print('Sample row 2:', file=sys.stderr)
        for i, cell in enumerate(ws[2], 1):
            print(f'  Col {i}: {cell.value}', file=sys.stderr)
    
    print('SUCCESS', file=sys.stderr)
except Exception as e:
    print(f'Error: {e}', file=sys.stderr)
    import traceback
    traceback.print_exc(file=sys.stderr)
